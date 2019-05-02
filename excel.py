#!/usr/bin/env python

'''
Contains functions for writing Excel files for the Standard Product Report
'''
from __future__ import print_function
import re
import json
import pickle
import hashlib
from openpyxl import Workbook
import dateutil.parser

def generate(aoi, track, acqs, slcs, acq_lists, ifg_cfgs, ifgs, audit_trail, enumeration=False):
    '''ingests the various products and stages them by track for generating worksheets'''
    # unique tracks based on acquisition list
    print('generating workbook for track {}'.format(track))
    generate_track(track, aoi, acqs, slcs, acq_lists, ifg_cfgs, ifgs, audit_trail, enumeration)

def generate_track(track, aoi, acqs, slcs, acq_lists, ifg_cfgs, ifgs, audit_trail, enumeration):
    '''generates excel sheet for given track, inputs are lists'''
    # stage products
    filename = '{}_T{}.xlsx'.format(aoi.get('_id', 'AOI'), track)
    acq_dct = convert_to_dict(acqs) # converts to dict based on id
    slc_dct = convert_to_dict(slcs) # converts to dict based on id
    acq_map = resolve_slcs_from_acqs(acqs) # converts acquisition ids to slc ids
    slc_map = resolve_acqs_from_slcs(acqs) # converts slc ids to acq_ids
    acq_list_dct = convert_to_hash_dict(acq_lists, conversion_dict=acq_map) # converts dict where key is hash of master/slave slc ids
    ifg_cfg_dct = convert_to_hash_dict(ifg_cfgs, conversion_dict=acq_map) # converts dict where key is hash of master/slave slc ids
    ifg_dct = convert_to_hash_dict(ifgs, conversion_dict=False) # converts dict where key is hash of master/slave slc ids
    
    # generate the acquisition sheet
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Enumerated Products"
    all_missing_slcs = [] # list of missing slcs by acquisition id
    titlerow = ['acquisition-list id', 'slcs localized?', 'ifg-cfg generated?', 'ifg generated?', 'missing slc ids', 'missing acq ids']
    ws1.append(titlerow)
    # for each acquisition list, determine relevant metrics
    for hkey in acq_list_dct.keys():
        obj = acq_list_dct.get(hkey)
        acqlistid = obj.get('_source', {}).get('id', 'No acquisition id found')
        missing_acq_str = ''
        missing_slc_str = ''
        missing_slcs = get_missing_slcs(obj, acq_map, slc_dct) # get list of any missing slc ids
        slcs_are_localized = False
        if not missing_slcs:
           slcs_are_localized = True
        if len(missing_slcs) > 0:
            slcs_are_localized = False
            all_missing_slcs.extend(missing_slcs) # add to master list for later
            missing_slc_str = ', '.join(missing_slcs)
            missing_acqs = [slc_map.get(x, 'id_not_found') for x in missing_slcs]
            missing_acq_str = ', '.join(missing_acqs)
        row = [acqlistid, slcs_are_localized, in_dict(hkey, ifg_cfg_dct), in_dict(hkey, ifg_dct), missing_slc_str, missing_acq_str]
        ws1.append(row)
    # generate missing slc list
    ws2 = wb.create_sheet("Missing SLCs")
    all_missing_slcs = sorted(list(set(all_missing_slcs)))
    title_row = ['slc id', 'acquisition id', 'starttime', 'endtime']
    ws2.append(title_row)
    for slc_id in all_missing_slcs:
        acq_id = slc_map.get(slc_id)
        acq_obj = acq_dct.get(acq_id, {})
        starttime = acq_obj.get('_source', {}).get('starttime', '-')
        endtime = acq_obj.get('_source', {}).get('endtime', '-')
        row = [slc_id, acq_id, starttime, endtime]
        ws2.append(row)
    #determine all date pairs that should be generated
    ws3 = wb.create_sheet('Enumerated Date Pairs')
    all_date_pairs = []
    title_row = ['expected date pairs']
    ws3.append(title_row)
    for key in acq_list_dct.keys():
        acq_list = acq_list_dct[key]
        st = dateutil.parser.parse(acq_list.get('_source').get('starttime')).strftime('%Y%m%d')
        et = dateutil.parser.parse(acq_list.get('_source').get('endtime')).strftime('%Y%m%d')
        ts = '{}-{}'.format(et, st)
        all_date_pairs.append(ts)
    for dt in sorted(list(set(all_date_pairs))):
        ws3.append([dt])
    #all acquisitions
    ws4 = wb.create_sheet('Acquisitions')
    title_row = ['acquisition_id', 'starttime', 'endtime', 'slc_id', 'ipf_version']
    ws4.append(title_row)
    for key in sorted(acq_dct.keys()):
        acq = acq_dct[key]
        acq_id = acq.get('_id', 'UNKNOWN')
        acq_st = acq.get('_source', {}).get('starttime', False)
        acq_et = acq.get('_source', {}).get('endtime', False)
        slc_id = acq.get('_source', {}).get('metadata', {}).get('identifier', False)
        ipf_version = acq.get('_source', {}).get('metadata', {}).get('processing_version', False)
        ws4.append([acq_id, acq_st, acq_et, slc_id, ipf_version])
    #all slcs
    ws5 = wb.create_sheet('Localized SLCs')
    title_row = ['slc_id', 'starttime', 'endtime']
    ws5.append(title_row)
    for key in sorted(slc_dct.keys()):
        slc = slc_dct[key]
        slc_id = slc.get('_id', 'UNKNOWN')
        slc_st = slc.get('_source', {}).get('starttime', False)
        slc_et = slc.get('_source', {}).get('endtime', False)
        ws5.append([slc_id, slc_st, slc_et])
    #all ifg_cfgs
    ws6 = wb.create_sheet('IFG CFGs')
    title_row = ['ifg-cfg id', 'starttime', 'endtime']
    ws6.append(title_row)
    for key in ifg_cfg_dct.keys():
        slc = ifg_cfg_dct[key]
        slc_id = slc.get('_id', 'UNKNOWN')
        slc_st = slc.get('_source', {}).get('starttime', False)
        slc_et = slc.get('_source', {}).get('endtime', False)
        ws6.append([slc_id, slc_st, slc_et])
    #all ifgs
    ws7 = wb.create_sheet('IFGs')
    title_row = ['ifg id', 'starttime', 'endtime', 'acq-list', 'ifg-cfg']
    ws7.append(title_row)
    for key in ifg_dct.keys():
        slc = ifg_dct[key]
        slc_id = slc.get('_id', 'UNKNOWN')
        slc_st = slc.get('_source', {}).get('starttime', False)
        slc_et = slc.get('_source', {}).get('endtime', False)
        #determine if the ifg-cfg and acq-list exists for the ifg
        ifg_cfg = ifg_cfg_dct.get(key, False)
        if ifg_cfg:
            ifg_cfg = ifg_cfg.get('_id', False)
        acq_list = acq_list_dct.get(key, False)
        if acq_list:
            acq_list = acq_list.get('_id', False)
        ws7.append([slc_id, slc_st, slc_et, acq_list, ifg_cfg])
    #audit trail
    ws8 = wb.create_sheet('Audit Trail')
    #just write all keys
    title_row = audit_trail[0].get('_source', {}).get('metadata', {}).keys()
    for x in ['union_geojson', 'context']:
        title_row.remove(x)
    ws8.append(title_row)
    for element in audit_trail:
        met = element.get('_source', {}).get('metadata', {})
        publish_row = []
        for key in title_row:
            val = met.get(key, '')
            if not isinstance(val, str):
                val = json.dumps(val)
            publish_row.append(val)
        ws8.append(publish_row)
    ws9 = wb.create_sheet('Acquisition-Lists')
    title_row = ['acq-list id', 'master_scenes', 'slave_scenes', 'master_orbit_file', 'slave_orbit_file']
    ws9.append(title_row)
    for element in acq_lists:
        acq_id = element.get('_id', 'UNKNOWN')
        master_scenes = ', '.join(element.get('_source', {}).get('metadata', {}).get('master_scenes', []))
        slave_scenes = ', '.join(element.get('_source', {}).get('metadata', {}).get('slave_scenes', []))
        master_orbit_file = element.get('_source', {}).get('metadata', {}).get('master_orbit_file', [])
        slave_orbit_file = element.get('_source', {}).get('metadata', {}).get('slave_orbit_file', [])
        ws9.append([acq_id, master_scenes, slave_scenes, master_orbit_file, slave_orbit_file])

    #if there is an enumeration, generate the appropriate pages
    if enumeration is False:
        wb.save(filename)
        return
    # print the human enumerated list
    ws10 = wb.create_sheet('Input Enumerated Date Pairs')
    title_row = ['date_pairs']
    ws10.append(title_row)
    for date in enumeration:
        ws10.append([date])
    #generate the list of human versus algorithm derived date pairs
    ws11 = wb.create_sheet('Enumeration Comparison')
    title_row = ['Unique Date Pair', 'In Input Enumeration?', 'In HySDS Enumeration?', 'Reason HySDS Skipped', 'Audit Comment', 'Reference Failure']
    ws11.append(title_row)
    alg_date_pairs = all_date_pairs
    human_date_pairs = enumeration
    total_date_pairs = sorted(list(set(alg_date_pairs + human_date_pairs)))
    comment_dict = build_audit_dict(audit_trail, 'comment')
    failure_dict = build_audit_dict(audit_trail, 'failure_reason')
    for date_pair in total_date_pairs:
        in_human_enumeration = False
        if date_pair in human_date_pairs:
            in_human_enumeration = True
        in_alg_enumeration = False
        if date_pair in alg_date_pairs:
            in_alg_enumeration = True
        comment = comment_dict.get(date_pair, '')
        failure_reason = failure_dict.get(date_pair, '')
        ref_failure = failure_dict.get(date_pair[:8], '')
        ws11.append([date_pair, in_human_enumeration, in_alg_enumeration, failure_reason, comment, ref_failure])
    wb.save(filename)
 

def build_audit_dict(audit_trail, field):
    '''builds a dict that goes by YMD-YMD as key which returns the metadata field desired'''
    obj_dict = {}
    for element in audit_trail:
        met = element.get('_source', {}).get('metadata', {})
        #st = dateutil.parser.parse(met.get('starttime'))
        #et = dateutil.parser.parse(met.get('endtime'))
        #st_str = dateutil.parser.parse(st)
        try:
            reference_date = dateutil.parser.parse(met.get('reference_date', False)).strftime('%Y%m%d')
        except:
            reference_date = '00000000'
        try:
            secondary_date = dateutil.parser.parse(met.get('secondary_date', False)).strftime('%Y%m%d')
        except:
            secondary_date = '00000000'
        dt_str = '{}-{}'.format(reference_date, secondary_date)
        field_result = met.get(field, '')
        if obj_dict.get(dt_str, '') == '':
            obj_dict[dt_str] = field_result
        if obj_dict.get(reference_date, '') == '':
            obj_dict[reference_date] = field_result
    return obj_dict

def in_dict(hsh, dct):
    '''returns true if the hash input is a key in the input dict'''
    rslt = dct.get(hsh, False)
    if rslt is False:
        return False
    return True

def get_scenes(obj, stype='master'):
    '''returns the master/reference, or slave/secondary scene list'''
    met = obj.get('_source', {}).get('metadata', {})
    if stype is 'master' or stype is 'reference':
        lst = met.get('master_scenes', [])
        if not lst:
            lst = met.get('reference_scenes', [])
    if stype is 'slave' or stype is 'secondary':
        lst = met.get('slave_scenes', [])
        if not lst:
            lst = met.get('secondary_scenes', [])
    if not isinstance(lst, list):
        raise Exception('obj not returning list type: {}'.format(obj.get('_id', False)))
    return lst

def store_by_hash(obj_list, conversion_dict=False):
    '''converts the list into a dict of objects where the keys are a hash of their master & slave slcs. if the entry
       is acquisitions, uses a conversion dict to convert to slc ids'''
    out_dict = {}
    for obj in obj_list:
        master = get_scenes(obj, stype='master')
        slave = get_scenes(obj, stype='slave')
        #
        #master = obj.get('_source', {}).get('metadata', {}).get('master_scenes', [])
        #if not master:
        #    master = obj.get('_source', {}).get('metadata', {}).get('reference_scenes', [])
        #slave = obj.get('_source', {}).get('metadata', {}).get('slave_scenes', [])
        #if not slave:
        #    slave = obj.get('_source', {}).get('metadata', {}).get('secondary_scenes', [])
        if conversion_dict:
            master = [conversion_dict.get(x, '') for x in master]
            slave = [conversion_dict.get(x, '') for x in slave] 
        master = pickle.dumps(sorted(master))
        slave = pickle.dumps(sorted(slave))
        hsh = '{}_{}'.format(hashlib.md5(master).hexdigest(), hashlib.md5(slave).hexdigest())
        out_dict[hsh] = obj
    return out_dict

def gen_hash(master_slcs,  slave_slcs):
    '''generates a hash from the input master & slave slcs. Same as used in the enumerator'''
    master_ids_str=""
    slave_ids_str=""
    for slc in sorted(master_slcs):
        print("get_ifg_hash : master slc : %s" %slc)
        if isinstance(slc, tuple) or isinstance(slc, list):
            slc = slc[0]
        if master_ids_str=="":
            master_ids_str= slc
        else:
            master_ids_str += " "+slc
    for slc in sorted(slave_slcs):
        print("get_ifg_hash: slave slc : %s" %slc)
        if isinstance(slc, tuple) or isinstance(slc, list):
            slc = slc[0]
        if slave_ids_str=="":
            slave_ids_str= slc
        else:
            slave_ids_str += " "+slc
    id_hash = hashlib.md5(json.dumps([
            master_ids_str,
            slave_ids_str
            ]).encode("utf8")).hexdigest()
    return id_hash

def is_covered(obj, slc_dct):
    '''returns True if the SLCs are in slc_dct, False otherwise'''
    master = get_scenes(obj, stype='master')
    slave = get_scenes(obj, stype='slave')
    slc_list = list(set(master + slave))
    for slc_id in slc_list:
        if slc_dct.get(slc_id, False) is False:
            return False
    return True

def get_missing_slcs(obj, acq_map, slc_dct):
    '''returns the slc ids enumerated in the object that are not contained in the slc dict'''
    master = get_scenes(obj, stype='master')
    slave = get_scenes(obj, stype='slave')
    acq_ids = list(set(master + slave))
    #convert the acquisition ids to slc ids
    slc_ids = [acq_map.get(x, 'slc_id_not_found') for x in acq_ids]
    #if the slc ids are not in the slc dict
    missing = []
    for slc_id in slc_ids:
        if slc_dct.get(slc_id, False) is False:
            missing.append(slc_id)
    return missing

def convert_to_dict(input_list):
    '''attempts to convert the input list to a dict where the keys are object_id'''
    out_dict = {}
    for obj in input_list:
        obj_id = obj.get('_source', {}).get('id', False)
        out_dict[obj_id] = obj
    return out_dict

def convert_to_dt_dict(input_list):
    '''attempts to convert the input list to a dict where the keys are object_id'''
    out_dict = {}
    for obj in input_list:
        starttime = parse_start_time(obj)
        out_dict[starttime] = obj
    return out_dict
    
def parse_start_time(obj):
    '''gets start time'''
    st = obj.get('_source', {}).get('starttime', False)
    return dateutil.parser.parse(st).strftime('%Y-%m-%dT%H:%M:%S')

def parse_from_fn(obj_string):
    '''parses starttime from filename string'''
    reg = '([1-2][0-9]{7}T[0-9]{6})'
    dt = dateutil.parser.parse(re.findall(reg, obj_string)[0])
    return dt.strftime('%Y-%m-%dT%H:%M:%S')

def parse_slc_id(obj):
    '''returns the slc identifier for the acquisition list product'''
    obj_type = obj.get('_source', {}).get('dataset', False)
    if obj_type == 'acquisition-S1-IW_SLC':
        return obj.get('_source', {}).get('metadata', {}).get('identifier')
    if obj_type == 'S1-IW_SLC':
        return obj.get('_source', {}).get('id')
    return 'no_id_found'

def resolve_acqs_from_slcs(acqs):
    '''returns a dict that takes in an SLC id and returns the associated acq id'''
    mapping_dict = {}
    for acq in acqs:
        #slc_id = parse_slc_id(acq)
        slc_id = acq.get('_source', {}).get('metadata', {}).get('identifier')
        acq_id = acq.get('_source').get('id')
        #print(slc_id, ':', acq_id)
        mapping_dict[slc_id] = acq_id
    return mapping_dict
        
def resolve_slcs_from_acqs(acqs):
    '''returns a dict that takes in an acq id and returns the associated slc id'''
    mapping_dict = {}
    for acq in acqs:
        slc_id = acq.get('_source', {}).get('metadata', {}).get('identifier')
        #slc_id = parse_slc_id(acq)
        acq_id = acq.get('_source', {}).get('id', False)
        mapping_dict[acq_id] = slc_id
        #print(acq_id, ':', slc_id)
    return mapping_dict
